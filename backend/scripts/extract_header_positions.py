import csv
import json
import os
import re
import sys
from dataclasses import asdict, dataclass, field
from typing import Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


SECTION_RE = re.compile(r"^[一二三四五六七八九十\d]+[、.]")
READ_HEADER_RE = re.compile(r"示值|测量|读数|响应时间|reading|measured|result", re.IGNORECASE)
HEADER_HINTS = (
    "标准值",
    "标称值",
    "示值",
    "测量",
    "平均",
    "误差",
    "技术要求",
    "结论",
    "判定",
    "重复性",
    "响应时间",
    "不确定度",
)


def p(msg: str = "") -> None:
    enc = getattr(sys.stdout, "encoding", None) or "utf-8"
    sys.stdout.write(str(msg).encode(enc, errors="replace").decode(enc, errors="replace") + "\n")


def clean_text(value) -> str:
    if value is None:
        return ""
    return str(value).replace("\r", " ").replace("\n", " ").strip()


def is_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


@dataclass(frozen=True)
class PrintScope:
    min_row: int
    max_row: int
    min_col: int
    max_col: int

    def contains(self, row: int, col: int) -> bool:
        return self.min_row <= row <= self.max_row and self.min_col <= col <= self.max_col


@dataclass(frozen=True)
class MergeMeta:
    min_row: int
    max_row: int
    min_col: int
    max_col: int


@dataclass
class Slot:
    label: str
    col_start: int
    col_end: int


@dataclass
class HeaderGroup:
    header: str
    row: int
    col_start: int
    col_end: int
    slots: List["Slot"] = field(default_factory=list)


@dataclass
class CellItem:
    row: int
    cell: str
    type: str
    value: str


@dataclass
class SlotRecord:
    slot: str
    slot_col_range: str
    items: List[CellItem]


@dataclass
class HeaderRecord:
    header: str
    header_row: int
    header_col_range: str
    slots: List[SlotRecord]


@dataclass
class SectionRecord:
    section: str
    row_range: List[int]
    data_row_start: int
    records: List[HeaderRecord]


class ExcelTemplateParser:
    def __init__(self, worksheet):
        self.ws = worksheet
        self.scope = self._detect_print_scope()
        self.merged_lookup = self._build_merged_lookup()

    def parse(self) -> Dict:
        sections = self._find_sections()
        result_sections: List[SectionRecord] = []

        for section_title, section_start, section_end in sections:
            groups = self._find_header_groups(section_start, section_end)
            if not groups:
                continue

            header_band_end = self._detect_header_band_end(groups, section_end)
            data_start = min(section_end, header_band_end + 1)
            section_col_start = min(group.col_start for group in groups)
            section_col_end = max(group.col_end for group in groups)

            records: List[HeaderRecord] = []
            for group in groups:
                group.slots = self._build_slots(group, header_band_end)
                slot_records: List[SlotRecord] = []
                for slot in group.slots:
                    items = self._extract_slot_items(
                        slot=slot,
                        data_start=data_start,
                        data_end=section_end,
                        section_col_start=section_col_start,
                        section_col_end=section_col_end,
                    )
                    slot_records.append(
                        SlotRecord(
                            slot=slot.label,
                            slot_col_range=self._col_range_text(slot.col_start, slot.col_end),
                            items=items,
                        )
                    )

                records.append(
                    HeaderRecord(
                        header=group.header,
                        header_row=group.row,
                        header_col_range=self._col_range_text(group.col_start, group.col_end),
                        slots=slot_records,
                    )
                )

            result_sections.append(
                SectionRecord(
                    section=section_title,
                    row_range=[section_start, section_end],
                    data_row_start=data_start,
                    records=records,
                )
            )

        return {
            "sheet": self.ws.title,
            "max_row": self.ws.max_row,
            "max_column": self.ws.max_column,
            "scope": asdict(self.scope),
            "sections": [asdict(section) for section in result_sections],
        }

    def _detect_print_scope(self) -> PrintScope:
        if self.ws.print_area:
            area = self.ws.print_area
            if isinstance(area, (list, tuple)):
                area = area[0]
            area_text = str(area)
            if "!" in area_text:
                area_text = area_text.split("!", 1)[1]
            area_text = area_text.replace("$", "")
            col1, row1, col2, row2 = range_boundaries(area_text)
            return PrintScope(min_row=row1, max_row=row2, min_col=col1, max_col=col2)
        return PrintScope(min_row=1, max_row=self.ws.max_row, min_col=1, max_col=self.ws.max_column)

    def _build_merged_lookup(self) -> Dict[str, MergeMeta]:
        lookup: Dict[str, MergeMeta] = {}
        for cell_range in self.ws.merged_cells.ranges:
            meta = MergeMeta(
                min_row=cell_range.min_row,
                max_row=cell_range.max_row,
                min_col=cell_range.min_col,
                max_col=cell_range.max_col,
            )
            for row in range(meta.min_row, meta.max_row + 1):
                for col in range(meta.min_col, meta.max_col + 1):
                    lookup[f"{get_column_letter(col)}{row}"] = meta
        return lookup

    def _address(self, row: int, col: int) -> str:
        return f"{get_column_letter(col)}{row}"

    def _top_left_of(self, row: int, col: int) -> Tuple[int, int]:
        meta = self.merged_lookup.get(self._address(row, col))
        if meta:
            return meta.min_row, meta.min_col
        return row, col

    def _merged_span_of(self, row: int, col: int) -> Tuple[int, int]:
        meta = self.merged_lookup.get(self._address(row, col))
        if meta:
            return meta.min_col, meta.max_col
        return col, col

    def _is_top_left(self, row: int, col: int) -> bool:
        top_row, top_col = self._top_left_of(row, col)
        return top_row == row and top_col == col

    def _logical_value(self, row: int, col: int):
        top_row, top_col = self._top_left_of(row, col)
        return self.ws.cell(top_row, top_col).value

    def _logical_text(self, row: int, col: int) -> str:
        return clean_text(self._logical_value(row, col))

    def _find_sections(self) -> List[Tuple[str, int, int]]:
        titles: List[Tuple[str, int]] = []
        note_row: Optional[int] = None
        for row in range(self.scope.min_row, self.scope.max_row + 1):
            title_text = self._logical_text(row, self.scope.min_col)
            if not title_text:
                continue
            if "备注" in title_text or "备 注" in title_text:
                note_row = row
            if SECTION_RE.search(title_text):
                titles.append((title_text, row))

        sections: List[Tuple[str, int, int]] = []
        for index, (title, start_row) in enumerate(titles):
            if index + 1 < len(titles):
                end_row = titles[index + 1][1] - 1
            else:
                end_row = note_row - 1 if note_row and note_row > start_row else self.scope.max_row
            sections.append((title, start_row, min(end_row, self.scope.max_row)))
        return sections

    def _find_header_groups(self, section_start: int, section_end: int) -> List[HeaderGroup]:
        scan_end = min(section_end, section_start + 3)
        groups: List[HeaderGroup] = []
        seen = set()

        for row in range(section_start + 1, scan_end + 1):
            is_primary_header_row = row == section_start + 1
            for col in range(self.scope.min_col, self.scope.max_col + 1):
                if not self._is_top_left(row, col):
                    continue
                text = self._logical_text(row, col)
                if not text or text.isdigit():
                    continue
                value = self._logical_value(row, col)
                if is_formula(value):
                    continue
                col_start, col_end = self._merged_span_of(row, col)
                if not self._looks_like_header(text, col_start, col_end, is_primary_header_row):
                    continue
                key = (text, row, col_start, col_end)
                if key in seen:
                    continue
                seen.add(key)
                groups.append(HeaderGroup(header=text, row=row, col_start=col_start, col_end=col_end))

        groups.sort(key=lambda item: (item.col_start, item.row))
        implicit_read_group = self._build_implicit_read_group(section_start, section_end, groups)
        if implicit_read_group:
            groups.append(implicit_read_group)
            groups.sort(key=lambda item: (item.col_start, item.row))
        return groups

    def _looks_like_header(self, text: str, col_start: int, col_end: int, is_primary_header_row: bool) -> bool:
        if any(hint in text for hint in HEADER_HINTS):
            return True
        if not is_primary_header_row:
            return False
        if col_end > col_start and len(text) >= 2:
            return True
        return len(text) > 12 or ":" in text or "：" in text

    def _build_implicit_read_group(
        self,
        section_start: int,
        section_end: int,
        groups: Sequence[HeaderGroup],
    ) -> Optional[HeaderGroup]:
        scan_end = min(section_end, section_start + 4)
        numeric_cells: List[Tuple[int, int, int, int]] = []
        for row in range(section_start + 1, scan_end + 1):
            for col in range(self.scope.min_col, self.scope.max_col + 1):
                if not self._is_top_left(row, col):
                    continue
                text = self._logical_text(row, col)
                if not text.isdigit():
                    continue
                col_start, col_end = self._merged_span_of(row, col)
                numeric_cells.append((row, col_start, col_end, int(text)))

        row_candidates: Dict[int, List[Tuple[int, int, int]]] = {}
        for row, col_start, col_end, number in numeric_cells:
            if 1 <= number <= 20:
                row_candidates.setdefault(row, []).append((col_start, col_end, number))

        for row, items in sorted(row_candidates.items()):
            numbers = sorted(item[2] for item in items)
            if len(numbers) < 2 or numbers[0] != 1:
                continue
            if any(numbers[index] - numbers[index - 1] != 1 for index in range(1, len(numbers))):
                continue
            start_col = min(item[0] for item in items)
            end_col = max(item[1] for item in items)
            if any(start_col >= group.col_start and end_col <= group.col_end for group in groups):
                continue
            return HeaderGroup(header="测量值", row=row, col_start=start_col, col_end=end_col)
        return None

    def _detect_header_band_end(self, groups: Sequence[HeaderGroup], section_end: int) -> int:
        candidate_end = max(group.row for group in groups)
        next_row = candidate_end + 1
        if next_row > section_end:
            return candidate_end
        if self._row_looks_like_subheader(next_row, min(group.col_start for group in groups), max(group.col_end for group in groups)):
            return next_row
        return candidate_end

    def _row_looks_like_subheader(self, row: int, col_start: int, col_end: int) -> bool:
        non_empty = 0
        for col in range(col_start, col_end + 1):
            top_row, top_col = self._top_left_of(row, col)
            if top_row != row or top_col != col:
                continue
            value = self.ws.cell(top_row, top_col).value
            if value is None:
                continue
            if is_formula(value):
                return False
            if clean_text(value):
                non_empty += 1
        return non_empty > 0

    def _build_slots(self, group: HeaderGroup, header_band_end: int) -> List[Slot]:
        slots: List[Slot] = []
        seen = set()
        read_like_header = bool(READ_HEADER_RE.search(group.header))

        for row in range(group.row + 1, header_band_end + 1):
            if not self._row_looks_like_subheader(row, group.col_start, group.col_end):
                continue
            for col in range(group.col_start, group.col_end + 1):
                if not self._is_top_left(row, col):
                    continue
                text = self._logical_text(row, col)
                if not text or text == group.header:
                    continue
                value = self._logical_value(row, col)
                if is_formula(value):
                    continue

                if text.isdigit():
                    number = int(text)
                    if group.header == "测量值":
                        pass
                    elif read_like_header and row == group.row + 1 and 1 <= number <= 10:
                        pass
                    else:
                        continue

                slot_col_start, slot_col_end = self._merged_span_of(row, col)
                key = (text, slot_col_start, slot_col_end)
                if key in seen:
                    continue
                seen.add(key)
                slots.append(Slot(label=text, col_start=slot_col_start, col_end=slot_col_end))

        if group.header == "测量值" and not slots:
            for col in range(group.col_start, group.col_end + 1):
                if not self._is_top_left(group.row, col):
                    continue
                text = self._logical_text(group.row, col)
                if not text.isdigit():
                    continue
                slot_col_start, slot_col_end = self._merged_span_of(group.row, col)
                slots.append(Slot(label=text, col_start=slot_col_start, col_end=slot_col_end))

        if slots:
            slots.sort(key=lambda item: item.col_start)
            return slots
        return [Slot(label=group.header, col_start=group.col_start, col_end=group.col_end)]

    def _row_has_data(self, row: int, col_start: int, col_end: int) -> bool:
        for col in range(col_start, col_end + 1):
            if not self._is_top_left(row, col):
                continue
            if clean_text(self._logical_value(row, col)):
                return True
        return False

    def _extract_slot_items(
        self,
        slot: Slot,
        data_start: int,
        data_end: int,
        section_col_start: int,
        section_col_end: int,
    ) -> List[CellItem]:
        items: List[CellItem] = []
        for row in range(max(data_start, self.scope.min_row), min(data_end, self.scope.max_row) + 1):
            if not self._row_has_data(row, section_col_start, section_col_end):
                continue
            seen_cells = set()
            for col in range(slot.col_start, slot.col_end + 1):
                top_row, top_col = self._top_left_of(row, col)
                if not self.scope.contains(top_row, top_col):
                    continue
                key = (top_row, top_col)
                if key in seen_cells:
                    continue
                seen_cells.add(key)

                value = self.ws.cell(top_row, top_col).value
                text = clean_text(value)
                if not text:
                    continue
                items.append(
                    CellItem(
                        row=top_row,
                        cell=self._address(top_row, top_col),
                        type="formula" if is_formula(value) else "value",
                        value=text,
                    )
                )
        return items

    def _col_range_text(self, col_start: int, col_end: int) -> str:
        return f"{get_column_letter(col_start)}:{get_column_letter(col_end)}"


def extract(worksheet) -> Dict:
    return ExcelTemplateParser(worksheet).parse()


def print_summary(data: Dict) -> None:
    p(f"工作表: {data['sheet']}  范围: A1:{get_column_letter(data['max_column'])}{data['max_row']}")
    scope = data.get("scope", {})
    p(
        "读取范围: "
        f"{get_column_letter(scope.get('min_col', 1))}{scope.get('min_row', 1)}:"
        f"{get_column_letter(scope.get('max_col', 1))}{scope.get('max_row', 1)}"
    )
    for section in data["sections"]:
        p("\n" + "=" * 80)
        p(f"区段: {section['section']}  行范围: {section['row_range'][0]}-{section['row_range'][1]}")
        p(f"数据起始行: {section['data_row_start']}")
        p("-" * 80)
        p(f"{'表头':<22} | {'子槽位':<10} | {'列范围':<8} | {'提取数量':<8} | 示例")
        p("-" * 80)
        for record in section["records"]:
            for slot in record["slots"]:
                sample = ""
                if slot["items"]:
                    first_item = slot["items"][0]
                    sample = f"{first_item['cell']}={first_item['value'][:18]}"
                p(
                    f"{record['header'][:22]:<22} | "
                    f"{slot['slot'][:10]:<10} | "
                    f"{slot['slot_col_range']:<8} | "
                    f"{len(slot['items']):<8} | {sample}"
                )


def export_csv(data: Dict, out_base_path: str) -> str:
    out_csv = out_base_path + ".extracted.csv"
    with open(out_csv, "w", newline="", encoding="utf-8-sig") as file_obj:
        writer = csv.writer(file_obj)
        writer.writerow(
            [
                "sheet",
                "section",
                "section_row_start",
                "section_row_end",
                "header",
                "header_row",
                "header_col_range",
                "slot",
                "slot_col_range",
                "item_index",
                "data_row",
                "cell",
                "type",
                "value",
            ]
        )

        for section in data.get("sections", []):
            for record in section.get("records", []):
                for slot in record.get("slots", []):
                    items = slot.get("items", [])
                    if not items:
                        writer.writerow(
                            [
                                data.get("sheet", ""),
                                section.get("section", ""),
                                section.get("row_range", [None, None])[0],
                                section.get("row_range", [None, None])[1],
                                record.get("header", ""),
                                record.get("header_row", ""),
                                record.get("header_col_range", ""),
                                slot.get("slot", ""),
                                slot.get("slot_col_range", ""),
                                "",
                                "",
                                "",
                                "",
                                "",
                            ]
                        )
                        continue

                    for index, item in enumerate(items, 1):
                        writer.writerow(
                            [
                                data.get("sheet", ""),
                                section.get("section", ""),
                                section.get("row_range", [None, None])[0],
                                section.get("row_range", [None, None])[1],
                                record.get("header", ""),
                                record.get("header_row", ""),
                                record.get("header_col_range", ""),
                                slot.get("slot", ""),
                                slot.get("slot_col_range", ""),
                                index,
                                item.get("row", ""),
                                item.get("cell", ""),
                                item.get("type", ""),
                                item.get("value", ""),
                            ]
                        )
    return out_csv


def export_json(data: Dict, out_base_path: str) -> str:
    out_json = out_base_path + ".extracted.json"
    with open(out_json, "w", encoding="utf-8") as file_obj:
        json.dump(data, file_obj, ensure_ascii=False, indent=2)
    return out_json


def main() -> None:
    if len(sys.argv) < 2:
        p("用法: py extract_header_positions.py <xlsx路径>")
        sys.exit(1)

    workbook_path = sys.argv[1]
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = workbook.active

    extracted = extract(worksheet)
    print_summary(extracted)

    base_path = os.path.splitext(workbook_path)[0]
    out_json = export_json(extracted, base_path)
    out_csv = export_csv(extracted, base_path)
    p("\n已输出结构化结果: " + out_json)
    p("已输出CSV结果: " + out_csv)


if __name__ == "__main__":
    main()
