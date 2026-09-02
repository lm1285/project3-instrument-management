import os
import sys
from typing import Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries


def p(*args, sep=" ", end="\n", file=None):
    out = file or sys.stdout
    msg = sep.join(str(a) for a in args)
    enc = getattr(out, "encoding", None) or "utf-8"
    safe = msg.encode(enc, errors="replace").decode(enc, errors="replace")
    out.write(safe + end)


def safe_text(v) -> str:
    return "" if v is None else str(v)


def build_merged_map(sheet) -> Dict[str, Tuple[int, int, int, int]]:
    mapping: Dict[str, Tuple[int, int, int, int]] = {}
    for m in sheet.merged_cells.ranges:
        meta = (m.min_row, m.max_row, m.min_col, m.max_col)
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                mapping[f"{get_column_letter(c)}{r}"] = meta
    return mapping


def get_value(sheet, row: int, col: int, merged_map) -> str:
    addr = f"{get_column_letter(col)}{row}"
    if addr in merged_map:
        min_row, _, min_col, _ = merged_map[addr]
        return safe_text(sheet.cell(row=min_row, column=min_col).value)
    return safe_text(sheet.cell(row=row, column=col).value)


def is_top_left(row: int, col: int, merged_map) -> bool:
    addr = f"{get_column_letter(col)}{row}"
    if addr not in merged_map:
        return True
    min_row, _, min_col, _ = merged_map[addr]
    return row == min_row and col == min_col


def print_merged_cells(sheet):
    merged = list(sheet.merged_cells.ranges)
    if not merged:
        p("\n[合并单元格] 无")
        return

    p(f"\n[合并单元格] 共 {len(merged)} 个")
    for i, m in enumerate(merged, 1):
        range_addr = f"{get_column_letter(m.min_col)}{m.min_row}:{get_column_letter(m.max_col)}{m.max_row}"
        v = safe_text(sheet.cell(row=m.min_row, column=m.min_col).value)
        if len(v) > 24:
            v = v[:24] + "..."
        p(f"  [{i:02d}] {range_addr:<12} 内容: {v or '(空)'}")


def find_title_cells(sheet, title: str, merged_map, header_scan_rows: int = 20) -> List[Tuple[int, int]]:
    hits: List[Tuple[int, int]] = []
    max_row = min(header_scan_rows, sheet.max_row)
    for r in range(1, max_row + 1):
        for c in range(1, sheet.max_column + 1):
            if not is_top_left(r, c, merged_map):
                continue
            val = get_value(sheet, r, c, merged_map)
            if title in val:
                hits.append((r, c))
    return hits


def resolve_col_range_of_title(sheet, row: int, col: int, merged_map) -> Tuple[int, int]:
    addr = f"{get_column_letter(col)}{row}"
    if addr in merged_map:
        _, _, min_col, max_col = merged_map[addr]
        return min_col, max_col
    return col, col


def region_between_two_titles(sheet, left_title: str, right_title: str, merged_map) -> Optional[Tuple[int, int, int, int]]:
    left_hits = find_title_cells(sheet, left_title, merged_map)
    right_hits = find_title_cells(sheet, right_title, merged_map)
    if not left_hits or not right_hits:
        return None

    left_row, left_col = left_hits[0]
    right_row, right_col = right_hits[0]

    left_start, left_end = resolve_col_range_of_title(sheet, left_row, left_col, merged_map)
    right_start, _ = resolve_col_range_of_title(sheet, right_row, right_col, merged_map)

    start_col = left_start
    end_col = max(left_end, right_start - 1)
    return left_row, right_row, start_col, end_col


def get_cell_position(row: int, col: int, merged_map) -> str:
    addr = f"{get_column_letter(col)}{row}"
    if addr in merged_map:
        min_row, max_row, min_col, max_col = merged_map[addr]
        if min_row == max_row and min_col == max_col:
            return addr
        return f"{get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}"
    return addr


def find_measurement_data_regions(sheet, merged_map, header_scan_rows: int = 25) -> List[Dict]:
    """识别测量值数据区域（不含表头、二级表头、平均值列）"""
    from typing import Dict as DictType
    results: List[DictType] = []
    
    measure_keywords = ["仪器示值", "实测报警值", "响应时间"]
    
    found_regions = set()
    
    for r in range(1, min(header_scan_rows, sheet.max_row) + 1):
        for c in range(1, sheet.max_column + 1):
            if not is_top_left(r, c, merged_map):
                continue
            
            val = get_value(sheet, r, c, merged_map)
            if not any(kw in val for kw in measure_keywords):
                continue
            
            if any(ch in val for ch in ["七、", "六、", "五、", "四、", "三、", "二、", "一、"]):
                continue
            
            start_col, end_col = resolve_col_range_of_title(sheet, r, c, merged_map)
            
            region_key = (r, start_col, end_col)
            if region_key in found_regions:
                continue
            
            sub_header_end = r
            avg_col_ranges = []
            
            for check_r in range(r + 1, min(r + 4, sheet.max_row) + 1):
                is_header_row = False
                header_count = 0
                seq_nums = []
                for check_c in range(start_col, end_col + 1):
                    if not is_top_left(check_r, check_c, merged_map):
                        continue
                    cell_val = get_value(sheet, check_r, check_c, merged_map).strip()
                    if cell_val:
                        if cell_val.isdigit() and int(cell_val) <= 15:
                            seq_nums.append(int(cell_val))
                        elif "平均" in cell_val or cell_val in ["AVG", "avg"]:
                            is_header_row = True
                            header_count += 1
                            avg_start, avg_end = resolve_col_range_of_title(sheet, check_r, check_c, merged_map)
                            avg_col_ranges.append((avg_start, avg_end))
                
                if seq_nums:
                    seq_nums.sort()
                    is_sequential = True
                    for i in range(1, len(seq_nums)):
                        if seq_nums[i] - seq_nums[i-1] != 1:
                            is_sequential = False
                            break
                    if is_sequential and len(seq_nums) >= 2:
                        is_header_row = True
                        header_count += len(seq_nums)
                
                if is_header_row and header_count >= 1:
                    sub_header_end = check_r
                else:
                    break
            
            data_start_row = sub_header_end + 1
            
            data_end_row = data_start_row
            for check_r in range(data_start_row, min(data_start_row + 15, sheet.max_row) + 1):
                has_data = False
                is_section_title = False
                for check_c in range(start_col, end_col + 1):
                    cell_val = get_value(sheet, check_r, check_c, merged_map).strip()
                    if cell_val:
                        if any(ch in cell_val for ch in ["七、", "六、", "五、", "四、", "三、", "二、", "一、"]):
                            is_section_title = True
                            break
                        if "备 注" not in cell_val and "备注" not in cell_val:
                            has_data = True
                
                if is_section_title:
                    break
                if has_data:
                    data_end_row = check_r
                else:
                    break
            
            if data_end_row >= data_start_row:
                measure_cols = list(range(start_col, end_col + 1))
                
                for avg_start, avg_end in avg_col_ranges:
                    for avg_c in range(avg_start, avg_end + 1):
                        if avg_c in measure_cols:
                            measure_cols.remove(avg_c)
                
                if measure_cols:
                    measure_start_col = min(measure_cols)
                    measure_end_col = max(measure_cols)
                    measure_region = f"{get_column_letter(measure_start_col)}{data_start_row}:{get_column_letter(measure_end_col)}{data_end_row}"
                    measure_cols_count = len(measure_cols)
                else:
                    measure_region = ""
                    measure_cols_count = 0
                
                avg_regions = []
                for avg_start, avg_end in avg_col_ranges:
                    avg_region = f"{get_column_letter(avg_start)}{data_start_row}:{get_column_letter(avg_end)}{data_end_row}"
                    avg_regions.append(avg_region)
                
                found_regions.add(region_key)
                results.append({
                    "title": val.strip(),
                    "title_pos": get_cell_position(r, c, merged_map),
                    "data_region": measure_region,
                    "avg_regions": avg_regions,
                    "start_row": data_start_row,
                    "end_row": data_end_row,
                    "start_col": start_col,
                    "end_col": end_col,
                    "rows": data_end_row - data_start_row + 1,
                    "cols": measure_cols_count
                })
    
    return results


def find_pure_number_header_regions(sheet, merged_map, header_scan_rows: int = 25) -> List[Dict]:
    """识别纯数字表头的测量值区域（如表头只有1,2,3,4的情况）"""
    from typing import Dict as DictType
    results: List[DictType] = []
    
    found_regions = set()
    
    for r in range(1, min(header_scan_rows, sheet.max_row) + 1):
        seq_cells = []
        for c in range(1, sheet.max_column + 1):
            if not is_top_left(r, c, merged_map):
                continue
            
            cell_val = get_value(sheet, r, c, merged_map).strip()
            if cell_val.isdigit() and 1 <= int(cell_val) <= 20:
                start_col, end_col = resolve_col_range_of_title(sheet, r, c, merged_map)
                seq_cells.append({
                    "num": int(cell_val),
                    "col": c,
                    "start_col": start_col,
                    "end_col": end_col
                })
        
        if len(seq_cells) < 2:
            continue
        
        seq_cells.sort(key=lambda x: x["num"])
        
        is_sequential = True
        for i in range(1, len(seq_cells)):
            if seq_cells[i]["num"] - seq_cells[i-1]["num"] != 1:
                is_sequential = False
                break
        
        if not is_sequential:
            continue
        
        total_start_col = seq_cells[0]["start_col"]
        total_end_col = seq_cells[-1]["end_col"]
        
        region_key = (r, total_start_col, total_end_col)
        if region_key in found_regions:
            continue
        
        data_start_row = r + 1
        data_end_row = data_start_row
        
        for check_r in range(data_start_row, min(data_start_row + 15, sheet.max_row) + 1):
            has_data = False
            is_section_title = False
            for check_c in range(total_start_col, total_end_col + 1):
                cell_val = get_value(sheet, check_r, check_c, merged_map).strip()
                if cell_val:
                    if any(ch in cell_val for ch in ["七、", "六、", "五、", "四、", "三、", "二、", "一、"]):
                        is_section_title = True
                        break
                    if "备 注" not in cell_val and "备注" not in cell_val:
                        has_data = True
            
            if is_section_title:
                break
            if has_data:
                data_end_row = check_r
            else:
                break
        
        if data_end_row >= data_start_row:
            measure_region = f"{get_column_letter(total_start_col)}{data_start_row}:{get_column_letter(total_end_col)}{data_end_row}"
            
            found_regions.add(region_key)
            results.append({
                "title": f"测量值（序号 {seq_cells[0]['num']}-{seq_cells[-1]['num']}）",
                "title_pos": f"{get_column_letter(total_start_col)}{r}:{get_column_letter(total_end_col)}{r}",
                "data_region": measure_region,
                "avg_regions": [],
                "start_row": data_start_row,
                "end_row": data_end_row,
                "start_col": total_start_col,
                "end_col": total_end_col,
                "rows": data_end_row - data_start_row + 1,
                "cols": total_end_col - total_start_col + 1
            })
    
    return results


def print_measurement_regions(sheet, merged_map):
    """打印测量值数据区域"""
    regions = find_measurement_data_regions(sheet, merged_map)
    pure_num_regions = find_pure_number_header_regions(sheet, merged_map)
    
    all_regions = regions + pure_num_regions
    
    seen_data_regions = set()
    unique_regions = []
    for reg in all_regions:
        if reg['data_region'] and reg['data_region'] not in seen_data_regions:
            seen_data_regions.add(reg['data_region'])
            unique_regions.append(reg)
    
    if not unique_regions:
        p("\n[测量值数据区域] 未识别到")
        return
    
    p(f"\n[测量值数据区域] 共识别 {len(unique_regions)} 个")
    p("-" * 70)
    for i, reg in enumerate(unique_regions, 1):
        p(f"  [{i}] {reg['title']}")
        p(f"      标题位置: {reg['title_pos']}")
        if reg['data_region']:
            p(f"      测量值区域: {reg['data_region']} ({reg['rows']}行 x {reg['cols']}列)")
        for avg_reg in reg.get('avg_regions', []):
            p(f"      平均值区域: {avg_reg}")
    p("-" * 70)


def print_non_empty_preview(sheet, merged_map, start_row=1, start_col=1, max_rows=30, max_cols=20, cols_per_block=8):
    p(f"\n[内容预览] 行列网格（按列分块，避免终端自动换行错位）")
    max_rows = min(max_rows, sheet.max_row)
    max_cols = min(max_cols, sheet.max_column)
    col = start_col
    while col <= max_cols:
        block_end = min(max_cols, col + cols_per_block - 1)
        p(f"\n  [列块] {get_column_letter(col)}:{get_column_letter(block_end)}")
        header = ["ROW"] + [get_column_letter(c) for c in range(col, block_end + 1)]
        p("\t".join(header))

        for r in range(start_row, max_rows + 1):
            row_out: List[str] = [str(r)]
            for c in range(col, block_end + 1):
                if not is_top_left(r, c, merged_map):
                    val = ""
                else:
                    raw_val = get_value(sheet, r, c, merged_map).replace("\r", " ").replace("\n", " ").strip()
                    pos = get_cell_position(r, c, merged_map)
                    val = f"{raw_val} [{pos}]" if raw_val else ""
                row_out.append(val)
            p("\t".join(row_out))
        col = block_end + 1


def detect_read_scope(sheet) -> Tuple[int, int, int, int]:
    min_row, max_row = 1, sheet.max_row
    min_col, max_col = 1, sheet.max_column

    if sheet.print_area:
        area = sheet.print_area
        if isinstance(area, (list, tuple)):
            area = area[0]
        area = str(area)
        if "!" in area:
            area = area.split("!", 1)[1]
        area = area.replace("$", "")
        if ":" in area:
            c1, r1, c2, r2 = range_boundaries(area)
            return r1, r2, c1, c2

    return min_row, max_row, min_col, max_col


def analyze_excel(file_path: str, left_title: Optional[str] = None, right_title: Optional[str] = None):
    p("=" * 72)
    p(f"文件: {os.path.basename(file_path)}")
    p("=" * 72)

    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet = wb.active

    p(f"工作表: {sheet.title}")
    p(f"数据范围: A1:{get_column_letter(sheet.max_column)}{sheet.max_row}")

    merged_map = build_merged_map(sheet)
    s_min_r, s_max_r, s_min_c, s_max_c = detect_read_scope(sheet)
    p(f"读取范围(打印区域): {get_column_letter(s_min_c)}{s_min_r}:{get_column_letter(s_max_c)}{s_max_r}")
    
    print_measurement_regions(sheet, merged_map)
    
    print_non_empty_preview(
        sheet,
        merged_map,
        start_row=s_min_r,
        start_col=s_min_c,
        max_rows=s_max_r,
        max_cols=s_max_c
    )

    if left_title and right_title:
        region = region_between_two_titles(sheet, left_title, right_title, merged_map)
        if region is None:
            p(f"\n[标题区间] 未找到标题：'{left_title}' 或 '{right_title}'")
        else:
            left_row, right_row, start_col, end_col = region
            p("\n[标题区间]")
            p(f"  左标题: {left_title} (行 {left_row})")
            p(f"  右标题: {right_title} (行 {right_row})")
            p(
                f"  区域列范围: {get_column_letter(start_col)}-{get_column_letter(end_col)} "
                f"(共 {end_col - start_col + 1} 列)"
            )


if __name__ == "__main__":
    if len(sys.argv) < 2:
        p("用法:")
        p("  python analyze_template.py <xlsx路径> [左标题] [右标题]")
        sys.exit(1)

    file_path = sys.argv[1]
    left = sys.argv[2] if len(sys.argv) >= 4 else None
    right = sys.argv[3] if len(sys.argv) >= 4 else None

    analyze_excel(file_path, left, right)
